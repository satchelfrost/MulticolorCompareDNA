import openpyxl as opx
import pyperclip
from openpyxl import Workbook
from openpyxl.styles import PatternFill
wb = Workbook(write_only=True)
ws = wb.create_sheet()


def parseSeq(lines,seqName):
    
    '''splits each column'''
    data = []
    for line in lines: data.append(line.split(' '))
    '''removes any spaces'''
    for i in range(len(data)):
        for j in range(data[i].count('')): data[i].remove('')
    '''deletes the numbers at beginning of column'''
    for i in range(len(data)): del data[i][0]
    '''creates a list of lists from dna sequence'''
    seqRows = []
    for i in range(len(data)):
        seqRow = []
        seqRow.append(seqName)
        for j in range(len(data[i])):
            for k in range(len(data[i][j])):
                seqRow.append(data[i][j][k])
        seqRows.append(seqRow)    
    return seqRows

seqs = int(input('How many DNA sequences do you want to compare? '))
saveFile = input('What do you want to name the spreadsheet? ')

'''masterList contains each sequence, and each sequence is
   broken into rows'''
masterList = []
'''reads files so they can be parsed'''
for i in range(seqs):
    print('What is the name of DNA sequence',i+1,end='? ')
    name = input('')
    file = open(name+'.txt')
    info = file.readlines()
    masterList.append(parseSeq(info,name))
    file.close()

'''sequence that contains the most rows is used for following loop'''
elems = []
for i in range(len(masterList)): elems.append(len(masterList[i]))
bigElem = elems.index(max(elems))
    
'''adds dna sequence to excel spreadsheet, 60 columns, x rows'''
for row in range(len(masterList[bigElem])):
    for seq in range(len(masterList)):
        try:
            ws.append(masterList[seq][row])
        except IndexError:
            ws.append([])
    ws.append([])
    
wb.save(saveFile+'.xlsx')

'''color match'''
match = input('Do you want to color match your sequence (y/n)? ')
if match == 'y':
    wb = opx.load_workbook(saveFile+'.xlsx')
    sheet = wb['Sheet']
    ws = wb.active


    red = 'FFFF0000'
    green = '0000FF00'
    blue = 'FF0000FF'

    greenFill = PatternFill(start_color=green,
                       end_color=green,
                       fill_type='solid')
    redFill = PatternFill(start_color=red,
                       end_color=red,
                       fill_type='solid')
    blueFill = PatternFill(start_color=blue,
                       end_color=blue,
                       fill_type='solid')


    ws['BK1'] = 'Matched'
    ws['BK1'].fill = greenFill
    ws['BK2'] = 'Unmatched'
    ws['BK2'].fill = blueFill

    lastRow = sheet.max_row + 1
    end = int(lastRow / (seqs+1))

    for section in range(end):
        startSec = (seqs+1)*section + 1
        endSec = (seqs+1)*section + (seqs+1)
        for col in range(2,62):
            bp = []
            for row in range(startSec,endSec):
                cell = sheet.cell(row=row,column=col).value
                bp.append(cell)
            if bp.count(bp[0]) == seqs:
                for row in range(startSec,endSec):
                    sheet.cell(row=row,column=col).fill = greenFill
            else:
                for row in range(startSec,endSec):
                    sheet.cell(row=row,column=col).fill = blueFill
    wb.save(saveFile+'.xlsx')



